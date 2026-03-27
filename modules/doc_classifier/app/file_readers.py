import csv
import io
import json
import zipfile
from pathlib import Path
from xml.etree import ElementTree

from openpyxl import load_workbook
from pypdf import PdfReader, PdfWriter


TEXT_EXTENSIONS = {
    ".txt",
    ".md",
    ".json",
    ".xml",
    ".html",
    ".htm",
    ".yaml",
    ".yml",
    ".csv",
}

TEXT_BASED_OFFICE_EXTENSIONS = {".docx", ".xlsx", ".xlsm"}


def read_document_as_text(path: Path, max_chars: int) -> tuple[str, str]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv(path, max_chars), "csv_text"
    if suffix in {".xlsx", ".xlsm"}:
        return _read_xlsx(path, max_chars), "xlsx_text"
    if suffix == ".docx":
        return _read_docx(path, max_chars), "docx_text"
    if suffix in TEXT_EXTENSIONS:
        return _read_plain_text(path, max_chars), "plain_text"
    if suffix == ".pdf":
        return _read_pdf_text(path, max_chars), "pdf_text"
    return _read_plain_text(path, max_chars), "best_effort_text"


def split_pdf_by_size(path: Path, temp_dir: Path, max_chunk_bytes: int) -> list[Path]:
    temp_dir.mkdir(parents=True, exist_ok=True)
    reader = PdfReader(str(path))
    if not reader.pages:
        return [path]

    chunks: list[Path] = []
    start_page = 0
    part_number = 1
    total_pages = len(reader.pages)

    while start_page < total_pages:
        end_page = start_page
        last_good_payload: bytes | None = None
        last_good_end = start_page
        while end_page < total_pages:
            payload = _render_pdf_chunk(reader, start_page, end_page)
            if len(payload) > max_chunk_bytes and end_page > start_page:
                break
            last_good_payload = payload
            last_good_end = end_page + 1
            if len(payload) > max_chunk_bytes:
                break
            end_page += 1

        if last_good_payload is None:
            last_good_payload = _render_pdf_chunk(reader, start_page, start_page)
            last_good_end = start_page + 1

        chunk_path = temp_dir / f"{path.stem}.part{part_number:03d}.pdf"
        chunk_path.write_bytes(last_good_payload)
        chunks.append(chunk_path)
        start_page = last_good_end
        part_number += 1

    return chunks


def extract_pdf_page_images(path: Path, temp_dir: Path, max_images: int = 8) -> list[Path]:
    temp_dir.mkdir(parents=True, exist_ok=True)
    try:
        reader = PdfReader(str(path))
    except Exception:
        return []
    results: list[Path] = []
    image_counter = 1
    for page_index, page in enumerate(reader.pages, start=1):
        try:
            page_images = list(page.images)
        except Exception:
            page_images = []
        for image in page_images:
            name = getattr(image, "name", f"page-{page_index}-{image_counter}.bin")
            suffix = Path(name).suffix or ".bin"
            target = temp_dir / f"{path.stem}.page{page_index:03d}.{image_counter:02d}{suffix}"
            try:
                target.write_bytes(image.data)
            except Exception:
                continue
            results.append(target)
            image_counter += 1
            if len(results) >= max_images:
                return results
    return results


def _render_pdf_chunk(reader: PdfReader, start_page: int, end_page: int) -> bytes:
    writer = PdfWriter()
    for index in range(start_page, end_page + 1):
        writer.add_page(reader.pages[index])
    buffer = io.BytesIO()
    writer.write(buffer)
    return buffer.getvalue()


def _read_plain_text(path: Path, max_chars: int) -> str:
    for encoding in ("utf-8", "cp1251", "latin-1"):
        try:
            text = path.read_text(encoding=encoding)
            return _truncate_text(text, max_chars)
        except (OSError, UnicodeDecodeError):
            continue
    return ""


def _read_csv(path: Path, max_chars: int) -> str:
    rows: list[str] = []
    try:
        with path.open("r", encoding="utf-8", newline="") as handle:
            reader = csv.reader(handle)
            for row in reader:
                rows.append(" | ".join(cell.strip() for cell in row))
    except UnicodeDecodeError:
        with path.open("r", encoding="cp1251", newline="") as handle:
            reader = csv.reader(handle)
            for row in reader:
                rows.append(" | ".join(cell.strip() for cell in row))
    except OSError:
        return ""
    return _truncate_text("\n".join(rows), max_chars)


def _read_xlsx(path: Path, max_chars: int) -> str:
    try:
        workbook = load_workbook(filename=path, read_only=True, data_only=True)
    except Exception:
        return ""
    fragments: list[str] = []
    for sheet in workbook.worksheets:
        fragments.append(f"[Лист] {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            values = [str(value).strip() for value in row if value not in (None, "")]
            if values:
                fragments.append(" | ".join(values))
    return _truncate_text("\n".join(fragments), max_chars)


def _read_docx(path: Path, max_chars: int) -> str:
    try:
        with zipfile.ZipFile(path) as archive:
            payload = archive.read("word/document.xml")
    except (OSError, KeyError, zipfile.BadZipFile):
        return ""
    try:
        root = ElementTree.fromstring(payload)
    except ElementTree.ParseError:
        return ""
    namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    paragraphs: list[str] = []
    for paragraph in root.findall(".//w:p", namespace):
        parts = [node.text or "" for node in paragraph.findall(".//w:t", namespace)]
        text = "".join(parts).strip()
        if text:
            paragraphs.append(text)
    return _truncate_text("\n".join(paragraphs), max_chars)


def _read_pdf_text(path: Path, max_chars: int) -> str:
    try:
        reader = PdfReader(str(path))
    except Exception:
        return ""
    parts: list[str] = []
    for page_number, page in enumerate(reader.pages, start=1):
        try:
            text = page.extract_text() or ""
        except Exception:
            text = ""
        if text.strip():
            parts.append(f"[Страница {page_number}]\n{text.strip()}")
        candidate = "\n\n".join(parts)
        if len(candidate) >= max_chars:
            return _truncate_text(candidate, max_chars)
    return _truncate_text("\n\n".join(parts), max_chars)


def serialize_processing_features(payload: dict) -> str:
    return json.dumps(payload, ensure_ascii=False, indent=2)


def _truncate_text(text: str, max_chars: int) -> str:
    normalized = "\n".join(line.rstrip() for line in text.splitlines()).strip()
    if len(normalized) <= max_chars:
        return normalized
    return normalized[: max_chars - 64].rstrip() + "\n...[обрезано по лимиту входного текста]..."
